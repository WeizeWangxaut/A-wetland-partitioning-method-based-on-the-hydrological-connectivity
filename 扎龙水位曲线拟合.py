###拟合水位
import numpy as np
import matplotlib.pyplot as plt
import openpyxl


# 定义x、y散点坐标
wb = openpyxl.load_workbook('D:/1资料/0417沼泽连通特征/zhalongclass_normal/NATU淹没拟合normal.xlsx')        # 目录改一下
wb2 = openpyxl.load_workbook('D:/1资料/0417沼泽连通特征/zhalongclass_normal/平期R.xlsx')
#for i in range(1, 7):
i = 1
sheet = wb['Sheet%d' % i]# 水位
sheet2 = wb2['Sheet%d' % i]
max_rows = sheet.max_row
max_columns = sheet.max_column

# 跟这个x拟合
n = 1

aa = 2   # 把这个点放在第2行
x = []
for c in range(4, max_columns+1):
    x.append(float(sheet.cell(aa, c).value))
x = np.array(x)


# 将第3行及第3行后的所有点与第2行的点拟合

for r in range(3, max_rows+1):
    y = []
    for c in range(4, max_columns+1):
        y.append(float(sheet.cell(r, c).value))
    y_list = np.array(y)
    # 用3次多项式拟合
    f1 = np.polyfit(x, y, 3)
    p1 = np.poly1d(f1)
    x1 = sorted(x)
    yvals_list = p1(x1)
    plot1 = plt.scatter(x, y_list, c='b', marker='o', label='original values')
    plot2 = plt.plot(x1, yvals_list,'r', label='polyfit values')


    # #################################拟合优度R^2的计算######################################
    def __sst(y_list):
        """
        计算SST(total sum of squares) 总平方和
        :param y_no_predicted: List[int] or array[int] 待拟合的y
        :return: 总平方和SST
        """
        y_mean = sum(y_list) / len(y_list)
        s_list = [(y - y_mean) ** 2 for y in y_list]
        sst = sum(s_list)
        return sst


    def __ssr(yvals_list, y_list):
        """
        计算SSR(regression sum of squares) 回归平方和
        :param y_fitting: List[int] or array[int]  拟合好的y值
        :param y_no_fitting: List[int] or array[int] 待拟合y值
        :return: 回归平方和SSR
        """
        y_mean = sum(y_list) / len(y_list)
        s_list = [(y - y_mean) ** 2 for y in yvals_list]
        ssr = sum(s_list)
        return ssr


    def __sse(yvals_lsit, y_list):
        """
        计算SSE(error sum of squares) 残差平方和
        :param y_fitting: List[int] or array[int] 拟合好的y值
        :param y_no_fitting: List[int] or array[int] 待拟合y值
        :return: 残差平方和SSE
        """
        s_list = [(yvals_lsit[i] - y_list[i]) ** 2 for i in range(len(yvals_lsit))]
        sse = sum(s_list)
        return sse


    def goodness_of_fit(yvals_list, y_lsit):
        """
        计算拟合优度R^2
        :param y_fitting: List[int] or array[int] 拟合好的y值
        :param y_no_fitting: List[int] or array[int] 待拟合y值
        :return: 拟合优度R^2
        """
        SSR = __ssr(yvals_list, y_lsit)
        SST = __sst(y_list)
        rr = SSR / SST
        return rr


    rr = goodness_of_fit(yvals_list, y_list)
    print(rr)
    plt.xlabel("%.2f"% rr)
    plt.ylabel('y')
    plt.legend(loc=4)  # 指定legend的位置右下角
    r2 = {"%.2f"% rr}
    m = sheet.cell(r, 1).value
    plt.title("%d"% m)
    plt.savefig('D:/1资料/0417沼泽连通特征/zhalongclass_normal/拟合曲线/%d.jpg' % m)
    plt.close()
    f1 = f1.tolist()
    for cc in range(1, 5):
        sheet2.cell(n+1, cc + 3).value = f1[cc - 1]
        sheet2.cell(n+1, cc + 4).value = rr
    n += 1



# # 绘图

wb2.save('D:/1资料/0417沼泽连通特征/zhalongclass_normal/平期R.xlsx')